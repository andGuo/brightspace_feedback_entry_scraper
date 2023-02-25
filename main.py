from typing import Dict
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver
import os
import time
import xlwings
from openpyxl import load_workbook
from dotenv import dotenv_values
config = dotenv_values(".env")

options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.add_experimental_option("detach", True)
options.add_argument('--no-sandbox')
options.add_argument("--window-size=1920,1080")
browser = webdriver.Chrome(options=options)


def main():
    ##### Check/Set the constants below this line #####
    BASE_URL = "https://brightspace.carleton.ca/d2l/home"
    GRADING_PAGE_URL = "https://brightspace.carleton.ca/d2l/lms/dropbox/admin/mark/folder_submissions_users.d2l?ou=131240&db=176868"
    # The PATH needs to end with a slash
    PATH_TO_FEEDBACK_SHEETS = '/Users/aguo/Dev/2022-2023/Winter/2401/2/graded/'
    # Relative path to classlist .xlsx
    PATH_TO_CLASSLIST = "./COMP2401A Intro to Systems Programming (LEC) Winter 2023_GradesExport_2023-02-25-08-07.xlsx"

    # Seconds to wait for the feedback input page to load
    SLEEP_TIME = 3 # increase this value it's failing a lot while inputting grades

    # Excel Cells
    FEEDBACK_CELL = 'B8'
    MAX_GRADE_CELL = 'C5'
    ACTUAL_GRADE_CELL = 'B5'
    STUDENT_NAME_CELL = 'B2'
    STUDENT_ID_CELL = 'B3'

    ##### Probably won't need to mess with anything after this line #####

    sid_sname_dict = get_student_names(PATH_TO_CLASSLIST)

    # These will probably break one day and may need to be updated before use
    email_field = (By.ID, 'userNameInput')
    password_field = (By.ID, 'passwordInput')
    login_button = (By.ID, 'submitButton')
    sid_search_bar = (  # Shadow dom below
        By.XPATH,  "//*[contains(@placeholder,'Search For…')]")
    open_feedback = (
        By.XPATH, "//a[starts-with(@title,'Evaluate ')] | //a[starts-with(@title,'Draft saved for ')]")
    # Shadow dom below
    fullscreen_button = 'return document.querySelector("d2l-consistent-evaluation").shadowRoot.querySelector("d2l-consistent-evaluation-page").shadowRoot.querySelector("#evaluation-template > div:nth-child(3) > consistent-evaluation-right-panel").shadowRoot.querySelector("div > d2l-consistent-evaluation-right-panel-feedback").shadowRoot.querySelector("d2l-consistent-evaluation-right-panel-block > d2l-htmleditor").shadowRoot.querySelector("div.d2l-htmleditor-label-flex-container > div > div.d2l-htmleditor-flex-container > div.d2l-htmleditor-toolbar-container > d2l-htmleditor-toolbar-full").shadowRoot.querySelector("div.d2l-htmleditor-toolbar-container.d2l-htmleditor-toolbar-overflowing.d2l-htmleditor-toolbar-chomping.d2l-htmleditor-toolbar-measured > div.d2l-htmleditor-toolbar-pinned-actions > d2l-htmleditor-button-toggle:nth-child(2)").shadowRoot.querySelector("button")'
    grade_input = 'return document.querySelector("d2l-consistent-evaluation").shadowRoot.querySelector("d2l-consistent-evaluation-page").shadowRoot.querySelector("#evaluation-template > div:nth-child(3) > consistent-evaluation-right-panel").shadowRoot.querySelector("div > d2l-consistent-evaluation-right-panel-grade-result").shadowRoot.querySelector("d2l-consistent-evaluation-right-panel-block > d2l-labs-d2l-grade-result-presentational").shadowRoot.querySelector("div.d2l-grade-result-presentational-container > d2l-grade-result-numeric-score").shadowRoot.querySelector("#grade-input")'
    save_as_draft_button = 'return document.querySelector("d2l-consistent-evaluation").shadowRoot.querySelector("d2l-consistent-evaluation-page").shadowRoot.querySelector("#evaluation-template > div:nth-child(4) > d2l-consistent-evaluation-footer").shadowRoot.querySelector("#consistent-evaluation-footer-save-draft")'

    # Gets login session token
    browser.get(BASE_URL)
    WebDriverWait(browser, 10).until(EC.element_to_be_clickable(
        email_field)).send_keys(config["USERNAME"])
    WebDriverWait(browser, 10).until(EC.element_to_be_clickable(
        password_field)).send_keys(config["PASSWORD"])
    WebDriverWait(browser, 10).until(
        EC.element_to_be_clickable(login_button)).click()

    # Gets all files in PATH_TO_FEEDBACK_SHEETS directory
    files = []
    for (_, _, filenames) in os.walk(PATH_TO_FEEDBACK_SHEETS):
        files.extend(filenames)

    # Opens every .xlsx file in PATH_TO_FEEDBACK_SHEETS and inputs feedback and grade
    for f in files:
        if f.endswith('.xlsx'):
            # this should probably be sanitised ¯\_(ツ)_/¯
            file_path = PATH_TO_FEEDBACK_SHEETS + f

            try:
                # hack to cache temp excel so that formulas are evaulated
                excel_app = xlwings.App(visible=False)
                excel_book = excel_app.books.open(file_path)
                excel_book.save()
                excel_book.close()
                excel_app.quit()

                workbook = load_workbook(
                    filename=file_path, data_only=True, read_only=True)
                sheet = workbook.active

                assignment = {"feedback": sheet[FEEDBACK_CELL].value, "max_grade": sheet[MAX_GRADE_CELL].value,
                              "actual_grade": sheet[ACTUAL_GRADE_CELL].value, "sname": sheet[STUDENT_NAME_CELL].value, "sid": sheet[STUDENT_ID_CELL].value}

                workbook.close()
            except Exception as e:
                print(f'ERROR >>> on file_path:{file_path}')
                print(e)
                continue

            # Display status
            print(
                f"\n{assignment['sname']:>18} - {assignment['actual_grade']:2d}/{assignment['max_grade']:2d}", end=" ")

            if str(assignment['sid']) not in sid_sname_dict:
                print("(Failed.)")
                print(
                    f"ERROR >>> Student - {assignment['sname']} ({assignment['sid']}) not found in classlist!")
                print(f"{file_path}")
                continue

            try:
                browser.get(GRADING_PAGE_URL)
                # Search for student id on brightspace page
                WebDriverWait(browser, 10).until(
                    EC.element_to_be_clickable(open_feedback))
                shadow_host = browser.find_element(
                    sid_search_bar[0], sid_search_bar[1])
                ac = ActionChains(browser)
                ac.move_to_element(shadow_host).click().perform()
                search = browser.switch_to.active_element
                search.send_keys(Keys.COMMAND, 'a')
                search.send_keys(Keys.DELETE)
                search.send_keys(sid_sname_dict[str(assignment['sid'])])
                search.send_keys(Keys.ENTER)

                # TODO:
                # 1. Probably should raise some exception if not exactly one student results from the search
                # 2. Maybe add some regex to validate the student name vs. the search result name

                # Goto Feedback Page
                edit = WebDriverWait(browser, 10).until(EC.element_to_be_clickable(
                    open_feedback)).click()

                # Input Grade
                # Ideally the line below would work but the sleep time will have to do
                # WebDriverWait(browser, 10).until( EC.invisibility_of_element_located((By.CLASS_NAME, "d2l-body-unresolved")))
                time.sleep(SLEEP_TIME)
                shadow_host = browser.execute_script(grade_input)
                ac.move_to_element(shadow_host).click().perform()
                grade = browser.switch_to.active_element
                grade.send_keys(Keys.COMMAND, 'a')
                grade.send_keys(Keys.DELETE)
                grade.send_keys(assignment['actual_grade'])

                fs_button = browser.execute_script(fullscreen_button)
                fs_button.click()
                feedback_text = browser.switch_to.active_element
                feedback_text.send_keys(Keys.COMMAND, 'a')
                feedback_text.send_keys(Keys.DELETE)
                feedback_text.send_keys(assignment['feedback'])
                fs_button.click()

                save_button = browser.execute_script(save_as_draft_button)
                ac.move_to_element(save_button).click().perform()

                print("(Done.)", end=" ")
            except Exception as e:
                print("(Failed.)")
                print(
                    f"ERROR >>> Unable to input feedback for student({assignment['sid']}): {sid_sname_dict[str(assignment['sid'])]}")
                print(f"{file_path}")
                # print(e)
                continue


def get_student_names(path: str) -> Dict[int, str]:
    try:
        workbook = load_workbook(
            filename=path, data_only=True, read_only=True)
        worksheet = workbook.active

        classlist_dict = {}
        num_rows = 0

        # Count number of non-empty rows
        for row in worksheet:
            if not all(col.value is None for col in row):
                num_rows += 1

        # Add to dict
        for row in range(1, num_rows + 1):
            key = worksheet.cell(row, 1).value
            value = worksheet.cell(row, 2).value + ' ' + \
                worksheet.cell(row, 3).value
            classlist_dict[key] = value

        workbook.close()

        return classlist_dict

    except Exception as e:
        print(f'ERROR >>> Unable to parse classlist at:{path}')
        raise Exception(e)


if __name__ == "__main__":
    main()
